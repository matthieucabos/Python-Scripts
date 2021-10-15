import sys
import os
import pyexcel as p

import re

__author__="CABOS Matthieu"
__date__=12/10/2021

# Penser a recuperer les fichiers sur tftp.srv-prive.icgm.fr:/var/lib/tftpboot/snoop
# os.system('scp mcabos@tftp.srv-prive.icgm.fr:/var/lib/tftpboot/snoop/* .')

switch_dict={
'balard-1D-1':'10.14.0.49',
'balard-1G-1':'10.14.0.51',
'balard-2D-1':'10.14.0.58',
'balard-2G-1':'10.14.0.60',
'balard-2H-1':'10.14.0.62',
'balard-3D-1':'10.14.0.67',
'balard-3G-1':'10.14.0.69',
'balard-3G-2':'10.14.0.70',
'balard-4C-1':'10.14.0.74',
'balard-4D-1':'10.14.0.76',
'balard-4G-1':'10.14.0.78',
'balard-4H-1':'10.14.0.80',
'balard-SRV':'10.14.0.20',
'balard-SRV-SUP':'10.14.0.21',
'balard-srv-cines':'10.14.0.30',
'balard-sup-cines':'10.14.0.31'
}

switch_dict2={
'10.14.0.49':'Balard-1D-1',
'10.14.0.51':'Balard-1G-1',
'10.14.0.58':'Balard-2D-1',
'10.14.0.60':'Balard-2G-1',
'10.14.0.62':'Balard-2H-1',
'10.14.0.67':'Balard-3D-1',
'10.14.0.69':'Balard-3G-1',
'10.14.0.70':'Balard-3G-2',
'10.14.0.74':'Balard-4C-1',
'10.14.0.76':'Balard-4D-1',
'10.14.0.78':'Balard-4G-1',
'10.14.0.80':'Balard-4H-1',
'10.14.0.20':'Balard-SRV',
'10.14.0.21':'Balard-SRV-SUP',
'10.14.0.30':'Balard-SRV-CINES',
'10.14.0.31':'Balard-SUP-CINES'
}

def build_ip_mac_dict(tftp_Content):
	"""
		Building Ip 2 @Mac dictionnarry from tftp boot server files (connected people).
		We are getting the full connected Users Mac => IP dictionnary using regular expression :

			* **[0-9a-z]{4}\\.[0-9a-z]{4}\\.[0-9a-z]{4}**  : Give us the MAC adress since the tftpboot files
			* **([0-9]\\/){2}[0-9]***  : Give us the Hardware Cisco Port Number since the tftpboot files 
			* **\\d+\\.\\d+\\.\\d+\\.\\d+** : Give us the IP Adress since the tftpboot files
	
		=============== ========== ===============================
		**Parameters**   **Type**   **Description**
		*tftp_Content*   string     The tftpboot file raw content
		=============== ========== ===============================

		:Returns: Dictionnary : The dictionnary with ip/mac correspondance 
	"""
	ip2mac={}
	MAC=""
	regex = r"[0-9a-z]{4}\.[0-9a-z]{4}\.[0-9a-z]{4}"
	regex2=r"([0-9]\/){2}[0-9]*"
	ip=re.compile(r'\d+\.\d+\.\d+\.\d+')
	for line in tftp_Content:
		matches = re.finditer(regex, line, re.MULTILINE)
		res_ip=ip.match(line)
		for matchNum, match in enumerate(matches, start=1):
			if (not res_ip == None):
				MAC=match.group()[0:2]+":"+match.group()[2:4]+":"+match.group()[5:7]+":"+match.group()[7:9]+":"+match.group()[10:12]+":"+match.group()[12:14]
		matches = re.finditer(regex2, line, re.MULTILINE)
		for matchNum, match in enumerate(matches, start=1):
			GiPort=str(match.group())
			if (not res_ip == None):
				ip2mac[MAC]=(res_ip.group(0),GiPort)
	return ip2mac

def get_content(switch_name):
	"""
		Get content from file since the switch_name argument.
		This function read the file and store informations into the return value.

		============== ============ =============================================
		**Parameters**   **Type**   **Description**
		*switch_name*  String       The exact switch_name from switch_dict keys
		============== ============ =============================================

		:Returns: String : The full Content of the file stored into a String Variable
	"""
	f=open(switch_name,'r')
	return f.readlines()

def write_in_tmp(ip_switch):
	"""
		Get SNMP informations and store it into the tmp file.

		=============== ========== ===========================================
		**Parameters**   **Type**   **Description**
		*ip_switch*      String     The exact IP adress of the current switch
		=============== ========== ===========================================

		:Returns: None
	"""
	os.system('snmpwalk -v 1 -c comaccess '+str(ip_switch)+':161 1.3.6.1.2.1.2.2.1.6 > tmp'+str(ip_switch))

def Get_switch_port_dict(ip_switch):
	"""	
		Read the tmp file containing SNMP informations and sort and store them into a Dictionnary with form :
		@Mac : Hardware Port Number

		=============== ========== ===========================================
		**Parameters**   **Type**   **Description**
		*ip_switch*      String     The exact IP adress of the current switch
		=============== ========== ===========================================

		:Returns: Dictionnary : The dictionnary associating a @mac to the hardware port number
	"""
	liste_addr=os.popen('cat tmp'+str(ip_switch))
	regex = r"([0-9a-z]{2}:){5}[0-9a-z]{2}"
	regex2=r"\.[0-9]+"
	Switch_port_dict={}

	for line in liste_addr.readlines():
		current_mac=""
		current_port=""
		matches = re.finditer(regex, line, re.MULTILINE)
		matches2 = re.finditer(regex2, line, re.MULTILINE)

		for matchNum, match in enumerate(matches, start=1):    # Looking for @MAC
			current_mac=str(match.group())
		for matchNum, match in enumerate(matches2, start=1):    # Looking for corresponding port number
			current_port=str(match.group())
			if current_mac :
				Switch_port_dict[current_mac]=current_port[1:]

	return Switch_port_dict

def Get_Port_and_GB(ip_switch,Final_dict):
	"""
		Populate the Final Dictionnary with Hardware Port Number values from Cisco SNMP Values (as verification of configuration...).

		=============== ============= ===========================================
		**Parameters**   **Type**     **Description**
		*ip_switch*      String       The exact IP adress of the current switch
		*Final_dict*     Dictionnary  The Final Dictionnary to be updated
		=============== ============= ===========================================

		:Returns: Dictionnary : The Final Dictionnary to be write updated
	"""
	Switch_port_dict=Get_switch_port_dict(ip_switch) # As example
	liste_addr=[]

	# Gettind @mac list
	for k in Switch_port_dict.keys():
		liste_addr.append(k)

	# Populate the Final Dictionnary to be write
	for k,v in Final_dict.items():
		if (v[0] in liste_addr):
			v[4]=Switch_port_dict[v[0]]
			v[5]=os.popen('./Cisco.sh '+str(v[3])+' 1 '+str(v[4])).read()
		pass

	return Final_dict

def Cisco2Socket(Cisco_name,*args):
	"""
		Getting the exact Room Socket Name from the GigabitEthernet Triolet provided by Cisco informations.

		=============== ========== =========================================================================================
		**Parameters**   **Type**   **Description**
		*Cisco_name*     String     The exact name of the Switch 
		*args*           String     A long string containing all the Hardware Cisco Port Number separated with a space key
		=============== ========== =========================================================================================

		:Returns: List : A List containing all the Room Socket Exact Name
	"""
	Socket_name=[]
	for i in range(len(args)):
		Socket_name.append(args[i])

	Cisco_list=[
	'Balard-EP-1',
	'Balard-PAC-1',
	'Balard-PAC-2',
	'Balard-RDC-1',
	'Balard-1C-1',
	'Balard-1D-1',
	'Balard-1G-1',
	'Balard-1G-2',
	'Balard-1H-1',
	'Balard-2C-1',
	'Balard-2D-1',
	'Balard-2G-1',
	'Balard-2H-1',
	'Balard-2H-2',
	'Balard-3C-1',
	'Balard-3D-1',
	'Balard-3G-1',
	'Balard-3G-2',
	'Balard-3H-1',
	'Balard-4C-1',
	'Balard-4D-1',
	'Balard-4G-1',
	'Balard-4H-1',
	'Balard-SRV',
	'Balard-SRV-SUP',
	'Balard-SRV-CINES',
	'Balard-SUP-CINES']
	
	f=open("Cisco2Socket.sh","a")
	f.write('#!/bin/bash\n# Author : CABOS Matthieu\n# Date : 08/10/2021\nterm shell\n')
	Cisco_Rep=[]	
	res={}

	for i in range(1,4):
		for j in range(1,49):
			f.write('show interface GigabitEthernet'+str(i)+'/0/'+str(j)+' | grep "N[0-9][A-Z][0-9][0-9]*-[0-9]*" \n')
	
	f.write('show interface GigabitEthernet0/0/0')
	f.close()
	os.system('ssh '+str(Cisco_name)+" < Cisco2Socket.sh > tmp2.txt")
	os.system('grep -v "^[[:space:]]*$" tmp2.txt > tmp2')
	os.system('rm tmp2.txt')
	i=7
	nb_ligne=int(os.popen('wc -l tmp2 | cut -d " " -f1').read())-i
	ind=1
	jnd=1
	while i <= nb_ligne:
		res[str(ind)+'/0/'+str(jnd)] = os.popen('cat tmp2 | head -'+str(i)+' | tail -2 | grep "N[0-9][A-Z][0-9][0-9]*-[0-9]*" | cut -d " " -f4 | sed "s/,//"').read()
		i+=2
		jnd+=1
		if jnd==49:
			ind=(ind + 1) if (ind <= 4) else 1 
			jnd=1
	os.system('rm tmp2')
	os.system('rm Cisco2Socket.sh')
	rez=[]
	GBname=''
	for socket in Socket_name :
		for k,v in res.items():
			if k==socket:
				GBname=v
				break		
		rez.append(GBname[:-1])
	return rez

def Cis2Socket(Cisco_name,Final_dict):
	"""
		Getting the exact Room Socket Name from the GigabitEthernet Triolet provided by Cisco informations.
		The Socket Name is stored in the given Dictionnary

		=============== ============= ==========================================
		**Parameters**   **Type**      **Description**
		*Cisco_name*     String        The exact name of the Switch 
		*Final_dict*     Dictionnary   The Final Dictionnary to be updated
		=============== ============= ==========================================

		:Returns: Dictionnary : The Final Dictionnary to be write updated
	"""
	regex=r'[A-Z][0-9][A-Z][0-9]+.[0-9]*'
	os.system('ssh -t '+str(Cisco_name)+' show interface description > Description_'+str(Cisco_name))
	f=open('Description_'+str(Cisco_name),'r')
	l=f.readlines()
	Tmp_dict={v:k for k,v in switch_dict2.items()}
	Plug_liste=[]

	for k,v in Final_dict.items():
		if v[4] == Tmp_dict[Cisco_name]:
			Plug_liste.append((v[6][2:],k))

	for line in l:
		for hw in Plug_liste:
		    if hw[0] in line:
		    	matches = re.finditer(regex, line,re.MULTILINE)
		    	for matchNum, match in enumerate(matches, start=1):
		    		tmp=Final_dict[hw[1]]
		    		tmp[7]=match.group()
		    		Final_dict[hw[1]]=tmp
		    	del Plug_liste[Plug_liste.index(hw)]

	return Final_dict


def update_Room_Sockets(ip_switch,Final_dict):
	"""
		Updating the Room Sockets Name field of the Dictionnary using the Cisco2Socket Procedure.
		Each Switch will be treated **independantly** from each others.
		It must be applied to each Switch to get the full Contents updated.

		=============== ============ =============================================
		**Parameters**   **Type**     **Description**
		*ip_switch*      String       The exact IP adress of the current switch
		*Final_dict*     Dictionnary  The Final Dictionnary to be updated
		=============== ============ =============================================

		:Returns: Dictionnary : The updated Dictionnary
	"""
	liste=[]
	Plug_liste=[]
	for k,v in Final_dict.items():
		if v[3]==ip_switch:
			liste.append(v[5])
	ind=0
	Plug_liste=Cisco2Socket(switch_dict2[ip_switch],' '.join(liste))	
	for k,v in Final_dict.items():
		if v[3]==ip_switch:
			try:
				v[6]=Plug_liste[ind]
				ind+=1
			except:
				break

	return Final_dict

def Get_Dpt(file_name):
	"""
		Getting Departement ID from the file_name ods file containing all the Vlans Informations.
		The Vlan name is readed and associated to its Id number.

		=============== =========== =========================
		**Parameters**   **Type**   **Description**
		*file_name*       String     An .ods file to read
		=============== =========== =========================

		:Returns: Dictionnary : The Departement dictionnary associating a Departement Id to a Computer Name on the network
	"""
	Dpt2Int_dict={
	'DPT1':510,
	'DPT2':511,
	'DPT3':512,
	'DPT4':513,
	'DPT5':514,
	'SGAF':524,
	'INSTRU-ON':515,
	'SSI':525,
	'INSTRU-OFF':516,
	'IMPRIM':518,
	'IDRAC-CIN':528,
	'IDRAC':501,
	'ExpProtect':526
	}

	Dpt_dict={}
	Records = p.get_array(file_name=file_name)
	Dpt_name=''

	for record in Records:
		for Dpt,v in Dpt2Int_dict.items():
			if Dpt in record[2] :
				Dpt_dict[record[0]]=v 
				break

	return Dpt_dict

def Get_Comm(file_name,Final_dict):
	"""
		Getting Comments fields from the '.ods' file_name.
		It returns a Dictionnary associating a Computer name to its Comments.
 
 		=============== ============= =============================================
		**Parameters**   **Type**      **Description**
		*file_name*      String         The .ods file_name to read
		*Final_dict*     Dictionnary    The Main Informations Dictionnary to read
		=============== ============= =============================================


		:Returns: Dictionnary : A Dictionnary associating Comments to the linked Computer Name on the network
	"""
	Comm_dict={}
	Records = p.get_array(file_name=file_name)
	Comm=''

	for record in Records:
		if record[0] in Final_dict.keys():
			Comm_dict[record[0]]=record[3]
	return Comm_dict

def Get_not_connected_dict(file_name,Final_dict):
	"""
		Similary building a Dictionnary with fewer informations for the disconnected Users.

		=============== ============== ===========================================
		**Parameters**   **Type**      **Description**
		*file_name*      String        The .ods file_name to read
		*Final_dict*     Dictionnary   The Main Informations Dictionnary to read
		=============== ============== ===========================================

		:Returns: Dictionnary :  A Dictionnary linking informations from the server to store the Disconnected Authorised Users
	"""
	Not_Conctd_Dict={}
	Records = p.get_array(file_name=file_name)
	dpt_dict=Get_Dpt(file_name)

	for record in Records:
		if not record[0] in Final_dict.keys():
			try:
				Not_Conctd_Dict[record[0]]=[record[1],Dpt_dict[record[0]],'','','','','','',record[3]]
			except:
				pass
	return Not_Conctd_Dict

# Getting infos from the Tftpboot server
os.system('scp mcabos@tftp.srv-prive.icgm.fr:/var/lib/tftpboot/snoop/* .')
Dpt_dict=Get_Dpt('../Ordinateurs.ods')

#Building IP:MAC dict
ip2mac={}
for switch in switch_dict.keys():
	Content=get_content(switch)
	ip2mac[switch]=build_ip_mac_dict(Content)

# Brownsing ods file
file_name='../Ordinateurs.ods'
records = p.get_array(file_name=file_name)
regex=r"/[0-9]+$"
Final_dict={}
Final_dict['Nom de la machine']=['@mac','Departement', '@ip machine', 'nom switch', '@ip switch', 'n° port', 'Triolet Gigabit','n° Prise','Commentaires']

# Searching current mac in @MAC database and updating Dictionnary Fields
for record in records:
	for switch in switch_dict.keys():
		for k,v in ip2mac[switch].items():
			if record[1] == k : 
				matches=re.finditer(regex,v[1],re.MULTILINE)
				for matchNum, match in enumerate(matches, start=1):
					port=match.group()[1:]
				Final_dict[record[0]]=[k,Dpt_dict[record[0]],v[0],switch,switch_dict[switch],port,"Gi"+v[1],"",'']

# Updating Comments Field
Comm=Get_Comm('../Ordinateurs.ods',Final_dict)
for k,v in Final_dict.items():
	if not (k == 'Nom de la machine'):
		tmp=v 
		tmp[8]=Comm[k]
		Final_dict[k]=tmp

# for sw in liste_switch:
# 	Final_dict=update_Room_Sockets(sw,Final_dict)

# Updating Room Sockets Names Field
for Cisco_name in switch_dict2.values():
	Final_dict=Cis2Socket(Cisco_name,Final_dict)

# Building the not-conected Dictionnarry
Not_Conctd_Dict=Get_not_connected_dict('../Ordinateurs.ods',Final_dict)

# Packaging as array to write
line=[]
to_write=[]
for k,v in Final_dict.items():
	line=[]
	line.append(k)
	line.extend(v)
	to_write.append(line)

to_write_ntc=[['Nom de la machine','@mac','Departement', '@ip machine', 'nom switch', '@ip switch', 'n° port', 'Triolet Gigabit','n° Prise','Commentaires']]
for k,v in Not_Conctd_Dict.items():
	line=[]
	line.append(k)
	line.extend(v)
	to_write_ntc.append(line)

Content={'Sheet 1':to_write, 'Sheet2':to_write_ntc}

# Saving ods file
book = p.Book(Content)
book.save_as('TftpBoot_List.xlsx')
os.system('rm Description*')
os.system('rm balard*')


# Setting Ods Document Layout
from openpyxl import *

Wb=load_workbook(filename='TftpBoot_List.xlsx')

border=styles.borders.Border(left=styles.borders.Side(style='medium'), 
                     right=styles.borders.Side(style='medium'), 
                     top=styles.borders.Side(style='medium'), 
                     bottom=styles.borders.Side(style='double'))
border2=styles.borders.Border(left=styles.borders.Side(style='thin'), 
                     right=styles.borders.Side(style='double'), 
                     top=styles.borders.Side(style='thin'), 
                     bottom=styles.borders.Side(style='thin'))
border3=styles.borders.Border(left=styles.borders.Side(style='thin'), 
                     right=styles.borders.Side(style='thin'), 
                     top=styles.borders.Side(style='thin'), 
                     bottom=styles.borders.Side(style='thin'))
font=styles.Font(color="00333333",size=12,bold=True)
font2=styles.Font(color="00333333",size=11,bold=False)
font3=styles.Font(color="00333300",italic=True)
fill = styles.PatternFill("solid",fgColor="DDDDDD")
fill2 = styles.PatternFill("solid",fgColor="e8e8e8")

for Ws in Wb.worksheets:
	for col in Ws.columns:
		maxi=0
		column=utils.get_column_letter(col[0].column)
		for cell in col:
			try:
				if(len(str(cell.value)) > maxi):
					maxi=len(cell.value)
			except:
				pass 
		adj_width=(maxi + 2)*1.2
		Ws.column_dimensions[column].width = adj_width
	Ws.showGridLines = True
	for i in range(1,11):
		Ws.cell(row=1,column=i).border=border
		Ws.cell(row=1,column=i).font=font
		Ws.cell(row=1,column=i).fill=fill
	for i in range(2,Ws.max_row+1):
		Ws.cell(row=i,column=1).border=border2
		Ws.cell(row=i,column=1).font=font2
		Ws.cell(row=i,column=1).fill=fill
		if(i<Ws.max_row):
			Ws.cell(row=i,column=2).font=font3
			Ws.cell(row=i,column=2).fill=fill2
			Ws.cell(row=i,column=2).border=border3
			Ws.cell(row=i,column=3).fill=fill2
			Ws.cell(row=i,column=3).border=border3
			Ws.cell(row=i,column=4).font=font3
			Ws.cell(row=i,column=4).fill=fill2
			Ws.cell(row=i,column=4).border=border3
			Ws.cell(row=i,column=5).fill=fill2
			Ws.cell(row=i,column=5).border=border3
			Ws.cell(row=i,column=6).font=font3
			Ws.cell(row=i,column=6).fill=fill2
			Ws.cell(row=i,column=6).border=border3
			Ws.cell(row=i,column=7).fill=fill2
			Ws.cell(row=i,column=7).border=border3
			Ws.cell(row=i,column=8).fill=fill2
			Ws.cell(row=i,column=8).border=border3
			Ws.cell(row=i,column=9).fill=fill2
			Ws.cell(row=i,column=9).border=border3
			Ws.cell(row=i,column=10).fill=fill2
			Ws.cell(row=i,column=10).border=border3
Wb.save(filename='TftpBoot_List.xlsx')

# Convert to .ods file
os.system('soffice --headless --convert-to ods *.xlsx')
os.system('rm *.xlsx')