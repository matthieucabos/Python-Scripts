import os
import sys
import pyexcel as p

__author__="CABOS Matthieu"
__date__=15/10/2021

# Update Ordinateurs.ods

Records = p.get_array(file_name='RESEAU.ods')

to_write=[]

Replace_dict={
'SGAF':'ICGM-SGAF',
'D3':'ICGM-DPT3-MPH',
'D4':'ICGM-DPT4-CMNME',
'SSI':'ICGM-SSI',
'Sous-réseau':'Sous-réseau',
'Didactique':'Didactique',
'':''
}
try:
	for record in Records:
		tmp=''
		for c in record[1].lower():
			if c != '-':
				tmp+=c
			else:
				tmp+=':'
		record[1]=tmp
		record[2]=Replace_dict[record[2]]
		record[3]=record[4]

	append=[]
	for record in Records:
		append.append(record[:4])
except:
	pass
Ordi=p.get_array(file_name='Ordinateurs.ods')
try:
	Ordi.extend(append[2:])
except:
	pass
Content={'Sheet 1':Ordi}
book = p.Book(Content)
book.save_as('UpdatedOrdinateurs.xlsx') 

from openpyxl import *

Wb=load_workbook(filename='UpdatedOrdinateurs.xlsx')

border=styles.borders.Border(left=styles.borders.Side(style='medium'), 
                     right=styles.borders.Side(style='medium'), 
                     top=styles.borders.Side(style='medium'), 
                     bottom=styles.borders.Side(style='double'))
border2=styles.borders.Border(left=styles.borders.Side(style='thin'), 
                     right=styles.borders.Side(style='thin'), 
                     top=styles.borders.Side(style='thin'), 
                     bottom=styles.borders.Side(style='thin'))
font=styles.Font(color="00333333",size=12,bold=True)
font2=styles.Font(color="00333300",italic=True,bold=True)
font3=styles.Font(color="00333300",italic=True)
fill = styles.PatternFill("solid",fgColor="DDDDDD")
fill2 = styles.PatternFill("solid",fgColor="e8e8e8")

for Ws in Wb.worksheets:
	for col in Ws.columns:
		maxi=0
		column=utils.get_column_letter(col[0].column)
		for cell in col:
			try:
				if(len(str(cell.value)) > maxi):
					maxi=len(cell.value)
			except:
				pass 
		adj_width=(maxi + 2)*1.2
		Ws.column_dimensions[column].width = adj_width
	Ws.showGridLines = True
	for i in range(1,5):
		Ws.cell(row=1,column=i).border=border
		Ws.cell(row=1,column=i).font=font
		Ws.cell(row=1,column=i).fill=fill
	for i in range(2,Ws.max_row+1):
		for j in range(1,5):
			Ws.cell(row=i,column=j).fill=fill2
			Ws.cell(row=i,column=j).border=border2
	for i in range(2,Ws.max_row+1):
		Ws.cell(row=i,column=1).font=font2
		Ws.cell(row=i,column=2).font=font3


Wb.save(filename='UpdatedOrdinateurs.xlsx')

# Convert to .ods file
os.system('soffice --headless --convert-to ods *.xlsx')
os.system('rm *.xlsx')